#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
leggi_schede.py
================
Legge le schede OMNIA-TFS scansionate (PDF multipagina o immagini) e produce
un file Excel con: risposte una riga per scheda, coda di revisione per i casi
dubbi, riepilogo con conteggi, percentuali e indice sintetico 1-5.

Federazione Nazionale Maestri del Lavoro - Settore Scuola

USO
---
  python leggi_schede.py scansione.pdf
  python leggi_schede.py scansione1.pdf scansione2.pdf --out risultati_3A.xlsx
  python leggi_schede.py cartella_immagini/*.jpg --template template.json

Il foglio Riepilogo e' tutto a formule: se correggi il foglio Dati dopo la
revisione manuale, conteggi e indici si aggiornano da soli.

DIPENDENZE
----------
  pip install opencv-python-headless numpy openpyxl pymupdf pillow
"""

import argparse
import glob
import json
import os
import sys
from datetime import datetime

import cv2
import numpy as np
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# --------------------------------------------------------------------------
# PARAMETRI DI RICONOSCIMENTO
# Ritoccare solo se la stampa o lo scanner cambiano in modo sensibile.
# --------------------------------------------------------------------------
DPI_RENDER = 200          # risoluzione di rasterizzazione del PDF
PX_PER_MM = 10.0          # risoluzione dell'immagine raddrizzata (254 dpi)
RAGGIO_CAMPIONE = 0.68    # frazione del raggio bolla campionata (esclude il bordo)
SOGLIA_PIENA = 0.38       # riempimento oltre il quale la casella e' annerita
SOGLIA_VUOTA = 0.16       # riempimento sotto il quale la casella e' vuota
SOGLIA_INCHIOSTRO_COMM = 0.004   # densita' minima per considerare scritti i commenti

STATO_OK = "OK"
STATO_REV = "DA VERIFICARE"

CAMPI_RIF = ["Consolato", "Scuola", "Classe", "Data", "Relatore"]


# ==========================================================================
# 1. ACQUISIZIONE IMMAGINI
# ==========================================================================
def pagine_da_file(percorso):
    """Restituisce una lista di (etichetta_pagina, immagine_grigia numpy)."""
    est = os.path.splitext(percorso)[1].lower()
    nome = os.path.basename(percorso)
    fuori = []

    if est == ".pdf":
        import pymupdf
        doc = pymupdf.open(percorso)
        for i, pagina in enumerate(doc):
            pix = pagina.get_pixmap(dpi=DPI_RENDER, colorspace=pymupdf.csGRAY)
            img = np.frombuffer(pix.samples, dtype=np.uint8).reshape(pix.height, pix.width)
            fuori.append(("%s p.%d" % (nome, i + 1), img.copy()))
        doc.close()
    else:
        img = cv2.imread(percorso, cv2.IMREAD_GRAYSCALE)
        if img is None:
            raise IOError("Impossibile leggere %s" % percorso)
        fuori.append((nome, img))
    return fuori


# ==========================================================================
# 2. RADDRIZZAMENTO SUI QUATTRO MARKER
# ==========================================================================
def trova_marker(gray):
    """Individua i quattro quadrati neri pieni agli angoli.

    Filtri applicati: area plausibile, forma quadrata, pieno (senza fori
    interni, cosi' i quadrati del QR code vengono scartati), posizione nella
    fascia esterna della pagina.
    """
    h, w = gray.shape
    lato_atteso = 6.0 / 210.0 * w          # 6 mm su una A4 larga 210 mm
    area_att = lato_atteso ** 2

    _, bw = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY_INV + cv2.THRESH_OTSU)
    bw = cv2.morphologyEx(bw, cv2.MORPH_CLOSE, np.ones((3, 3), np.uint8))

    contorni, gerarchia = cv2.findContours(bw, cv2.RETR_CCOMP, cv2.CHAIN_APPROX_SIMPLE)
    if gerarchia is None:
        return None
    gerarchia = gerarchia[0]

    candidati = []
    for i, cnt in enumerate(contorni):
        if gerarchia[i][2] != -1:          # ha figli -> non e' un quadrato pieno
            continue
        area = cv2.contourArea(cnt)
        if not (0.30 * area_att < area < 3.0 * area_att):
            continue
        x, y, cw, ch = cv2.boundingRect(cnt)
        if ch == 0:
            continue
        rapporto = cw / float(ch)
        if not (0.65 < rapporto < 1.55):
            continue
        if area / float(cw * ch) < 0.80:   # deve riempire il suo rettangolo
            continue
        cx, cy = x + cw / 2.0, y + ch / 2.0
        # solo la fascia esterna: 22% della pagina su entrambi gli assi
        if not ((cx < 0.22 * w or cx > 0.78 * w) and (cy < 0.22 * h or cy > 0.78 * h)):
            continue
        candidati.append((cx, cy, area))

    if len(candidati) < 4:
        return None

    angoli = [(0, 0), (w, 0), (w, h), (0, h)]
    scelti, usati = [], set()
    for ax, ay in angoli:
        migliore, dist_min = None, None
        for k, (cx, cy, _a) in enumerate(candidati):
            if k in usati:
                continue
            d = (cx - ax) ** 2 + (cy - ay) ** 2
            if dist_min is None or d < dist_min:
                dist_min, migliore = d, k
        if migliore is None:
            return None
        usati.add(migliore)
        scelti.append((candidati[migliore][0], candidati[migliore][1]))

    return np.array(scelti, dtype=np.float32)   # ordine: TL, TR, BR, BL


def raddrizza(gray, marker, larghezza_mm, altezza_mm):
    """Trasformazione prospettica sul rettangolo dei marker."""
    W = int(round(larghezza_mm * PX_PER_MM))
    H = int(round(altezza_mm * PX_PER_MM))
    dest = np.array([[0, 0], [W - 1, 0], [W - 1, H - 1], [0, H - 1]], dtype=np.float32)
    M = cv2.getPerspectiveTransform(marker, dest)
    return cv2.warpPerspective(gray, M, (W, H), flags=cv2.INTER_AREA,
                               borderValue=255)


# ==========================================================================
# 3. LETTURA DELLE CASELLE
# ==========================================================================
def maschera_inchiostro(warp):
    """Binarizza l'immagine raddrizzata: 255 dove c'e' inchiostro.

    Lo sfondo (la carta) viene stimato con una dilatazione a kernel largo:
    il kernel deve essere piu' grande della casella piena, altrimenti
    l'interno di una casella annerita verrebbe scambiato per carta.
    """
    k = int(round(6.0 * PX_PER_MM)) | 1               # 6 mm, dispari
    sfondo = cv2.dilate(warp, cv2.getStructuringElement(cv2.MORPH_RECT, (k, k)))
    sfondo = cv2.GaussianBlur(sfondo, (0, 0), PX_PER_MM)
    normalizzato = cv2.divide(warp, sfondo, scale=255)
    _, mask = cv2.threshold(normalizzato, 0, 255,
                            cv2.THRESH_BINARY_INV + cv2.THRESH_OTSU)
    return mask


def riempimento(mask, u, v, raggio_px):
    """Frazione di pixel anneriti dentro il disco interno della casella."""
    H, W = mask.shape
    cx, cy = int(round(u * W)), int(round(v * H))
    r = max(2, int(round(raggio_px * RAGGIO_CAMPIONE)))
    x0, x1 = max(0, cx - r), min(W, cx + r + 1)
    y0, y1 = max(0, cy - r), min(H, cy + r + 1)
    if x1 <= x0 or y1 <= y0:
        return 0.0
    ritaglio = mask[y0:y1, x0:x1]
    yy, xx = np.ogrid[y0:y1, x0:x1]
    disco = (xx - cx) ** 2 + (yy - cy) ** 2 <= r * r
    if disco.sum() == 0:
        return 0.0
    return float((ritaglio[disco] > 0).sum()) / float(disco.sum())


def leggi_domanda(mask, dom, raggio_px):
    """Ritorna (lista_indici_scelti, stato, dettaglio_riempimenti)."""
    valori = [riempimento(mask, o["u"], o["v"], raggio_px) for o in dom["opzioni"]]
    piene = [i for i, x in enumerate(valori) if x >= SOGLIA_PIENA]
    dubbie = [i for i, x in enumerate(valori) if SOGLIA_VUOTA <= x < SOGLIA_PIENA]

    stato = STATO_OK
    if dom["tipo"] == "singola":
        if len(piene) == 1 and not dubbie:
            pass
        elif len(piene) == 1 and dubbie:
            stato = "segno incerto su altra casella"
        elif len(piene) == 0:
            stato = "nessuna risposta"
        else:
            stato = "risposte multiple su domanda a scelta singola"
    else:
        if len(piene) == 0:
            stato = "nessuna risposta"
        elif dubbie:
            stato = "segno incerto su altra casella"

    return piene, stato, valori


# ==========================================================================
# 4. QR CODE E COMMENTI
# ==========================================================================
def leggi_qr(warp, zona):
    det = cv2.QRCodeDetector()
    H, W = warp.shape
    tentativi = []
    if zona:
        x0 = max(0, int(zona["u0"] * W)); x1 = min(W, int(zona["u1"] * W))
        y0 = max(0, int(zona["v0"] * H)); y1 = min(H, int(zona["v1"] * H))
        if x1 > x0 and y1 > y0:
            tentativi.append(warp[y0:y1, x0:x1])
    tentativi.append(warp)

    for img in tentativi:
        for fattore in (1, 2, 3):
            prova = img if fattore == 1 else cv2.resize(
                img, None, fx=fattore, fy=fattore, interpolation=cv2.INTER_CUBIC)
            try:
                testo, _pts, _q = det.detectAndDecode(
                    cv2.cvtColor(prova, cv2.COLOR_GRAY2BGR))
            except cv2.error:
                testo = ""
            if testo:
                return testo
    return ""


def scomponi_riferimento(payload):
    """'OMNIA|consolato|scuola|classe|data|relatore' -> dizionario."""
    vuoto = {c: "" for c in CAMPI_RIF}
    if not payload:
        return vuoto
    parti = payload.split("|")
    if parti and parti[0].upper() == "OMNIA":
        parti = parti[1:]
    for i, campo in enumerate(CAMPI_RIF):
        vuoto[campo] = parti[i].strip() if i < len(parti) else ""
    return vuoto


def ritaglia_commenti(warp, mask, zona, cartella, nome_base):
    """Salva il ritaglio dei commenti solo se contiene scrittura."""
    if not zona:
        return ""
    H, W = warp.shape
    x0 = max(0, int(zona["u0"] * W)); x1 = min(W, int(zona["u1"] * W))
    y0 = max(0, int(zona["v0"] * H)); y1 = min(H, int(zona["v1"] * H))
    if x1 <= x0 or y1 <= y0:
        return ""
    # margine interno per escludere il bordo stampato del riquadro
    m = int(2 * PX_PER_MM)
    xa, xb = min(x0 + m, x1 - 1), max(x1 - m, x0 + 1)
    ya, yb = min(y0 + int(6 * PX_PER_MM), y1 - 1), max(y1 - m, y0 + 1)
    if xb <= xa or yb <= ya:
        return ""
    densita = float((mask[ya:yb, xa:xb] > 0).mean())
    if densita < SOGLIA_INCHIOSTRO_COMM:
        return ""
    os.makedirs(cartella, exist_ok=True)
    percorso = os.path.join(cartella, "%s_commento.png" % nome_base)
    cv2.imwrite(percorso, warp[y0:y1, x0:x1])
    return percorso


# ==========================================================================
# 5. ELABORAZIONE DI UNA PAGINA
# ==========================================================================
def elabora_pagina(etichetta, gray, template, cartella_commenti, progressivo):
    rett = template["rettangolo_marker_mm"]
    esito = {
        "pagina": etichetta,
        "stato": STATO_OK,
        "note": [],
        "riferimento": {c: "" for c in CAMPI_RIF},
        "payload": "",
        "risposte": {},        # id_domanda -> lista indici
        "riempimenti": {},     # id_domanda -> lista valori (diagnostica)
        "commento_img": "",
    }

    marker = trova_marker(gray)
    if marker is None:
        esito["stato"] = STATO_REV
        esito["note"].append("marker non individuati: pagina bianca, storta o tagliata")
        return esito

    warp = raddrizza(gray, marker, rett["larghezza"], rett["altezza"])
    mask = maschera_inchiostro(warp)
    raggio_px = template["raggio_bolla_mm"] * PX_PER_MM

    payload = leggi_qr(warp, template.get("qr_zona"))
    esito["payload"] = payload
    esito["riferimento"] = scomponi_riferimento(payload)
    if not payload:
        esito["stato"] = STATO_REV
        esito["note"].append("QR non letto: riferimento da inserire a mano")

    for dom in template["domande"]:
        scelti, stato, valori = leggi_domanda(mask, dom, raggio_px)
        esito["risposte"][dom["id"]] = scelti
        esito["riempimenti"][dom["id"]] = [round(v, 3) for v in valori]
        if stato != STATO_OK:
            esito["stato"] = STATO_REV
            esito["note"].append("D%02d: %s" % (dom["numero"], stato))

    esito["commento_img"] = ritaglia_commenti(
        warp, mask, template.get("commenti_zona"), cartella_commenti,
        "scheda_%04d" % progressivo)
    return esito


# ==========================================================================
# 6. SCRITTURA DEL FILE EXCEL
# ==========================================================================
FONT = "Arial"
INTEST = PatternFill("solid", fgColor="1F6FB2")
GIALLO = PatternFill("solid", fgColor="FFF2CC")
BORDO = Border(*[Side(style="thin", color="BFBFBF")] * 4)


def _intesta(ws, valori, riga=1):
    for j, v in enumerate(valori, start=1):
        c = ws.cell(row=riga, column=j, value=v)
        c.font = Font(name=FONT, bold=True, color="FFFFFF", size=10)
        c.fill = INTEST
        c.alignment = Alignment(vertical="center", wrap_text=True)
    ws.freeze_panes = ws.cell(row=riga + 1, column=1)


def _larghezze(ws, larghezze):
    for j, w in enumerate(larghezze, start=1):
        ws.column_dimensions[get_column_letter(j)].width = w


def scrivi_excel(esiti, template, percorso):
    wb = Workbook()

    domande = template["domande"]
    scala_domande = [d for d in domande if d["scala"]]
    ha_commenti = template.get("commenti_zona") is not None

    # ---------------- foglio Risposte (leggibile) ----------------
    ws = wb.active
    ws.title = "Risposte"
    intest = ["N.", "Pagina"] + CAMPI_RIF + \
             ["D%02d - %s" % (d["numero"], d["testo"][:45]) for d in domande] + \
             ["Stato", "Note"]
    if ha_commenti:
        intest.append("Commento (immagine)")
    col_stato = intest.index("Stato") + 1
    _intesta(ws, intest)

    for i, e in enumerate(esiti, start=1):
        riga = [i, e["pagina"]] + [e["riferimento"][c] for c in CAMPI_RIF]
        for d in domande:
            scelti = e["risposte"].get(d["id"], [])
            riga.append("; ".join(d["opzioni"][k]["etichetta"] for k in scelti))
        riga += [e["stato"], " | ".join(e["note"])]
        if ha_commenti:
            riga.append(e["commento_img"])
        for j, v in enumerate(riga, start=1):
            c = ws.cell(row=i + 1, column=j, value=v)
            c.font = Font(name=FONT, size=10)
            c.border = BORDO
            c.alignment = Alignment(vertical="top", wrap_text=False)
        if e["stato"] != STATO_OK:
            ws.cell(row=i + 1, column=col_stato).fill = GIALLO
    _larghezze(ws, [5, 22] + [16] * len(CAMPI_RIF) + [26] * len(domande) +
               ([15, 40, 30] if ha_commenti else [15, 40]))

    # ---------------- foglio Dati (matrice 0/1, base dei calcoli) ----------------
    wd = wb.create_sheet("Dati")
    col_di = {}                     # (id_domanda, indice_opzione) -> lettera colonna
    intest_d = ["N."]
    for d in domande:
        for o in d["opzioni"]:
            intest_d.append("%s_%d" % (d["id"], o["indice"]))
            col_di[(d["id"], o["indice"])] = get_column_letter(len(intest_d))
    col_punteggio = {}
    for d in scala_domande:
        intest_d.append("%s_punteggio" % d["id"])
        col_punteggio[d["id"]] = get_column_letter(len(intest_d))
    intest_d.append("Valida")
    col_valida = get_column_letter(len(intest_d))
    _intesta(wd, intest_d)

    for i, e in enumerate(esiti, start=1):
        wd.cell(row=i + 1, column=1, value=i)
        for d in domande:
            scelti = set(e["risposte"].get(d["id"], []))
            for o in d["opzioni"]:
                wd.cell(row=i + 1,
                        column=_col_index(col_di[(d["id"], o["indice"])]),
                        value=1 if o["indice"] in scelti else 0)
        for d in scala_domande:
            scelti = e["risposte"].get(d["id"], [])
            val = d["scala"][scelti[0]] if len(scelti) == 1 else None
            wd.cell(row=i + 1, column=_col_index(col_punteggio[d["id"]]), value=val)
        wd.cell(row=i + 1, column=_col_index(col_valida),
                value=1 if e["stato"] == STATO_OK else 0)
    for riga in wd.iter_rows(min_row=2):
        for c in riga:
            c.font = Font(name=FONT, size=9)

    ultima = len(esiti) + 1
    wd.sheet_state = "hidden" if esiti else "visible"

    # ---------------- foglio Riepilogo (tutto a formule) ----------------
    wr = wb.create_sheet("Riepilogo", 1)
    wr["A1"] = "RIEPILOGO VALUTAZIONI \u2013 OMNIA TFS"
    wr["A1"].font = Font(name=FONT, bold=True, size=14, color="1F6FB2")
    wr["A2"] = "Schede elaborate:"
    wr["B2"] = len(esiti)
    wr["A3"] = "Schede senza segnalazioni:"
    wr["B3"] = "=SUM(Dati!%s2:%s%d)" % (col_valida, col_valida, max(2, ultima))
    wr["A4"] = "Generato il:"
    wr["B4"] = datetime.now().strftime("%d/%m/%Y %H:%M")
    for r in range(2, 5):
        wr.cell(row=r, column=1).font = Font(name=FONT, bold=True, size=10)
        wr.cell(row=r, column=2).font = Font(name=FONT, size=10)

    r = 6
    for d in domande:
        wr.cell(row=r, column=1, value="D%02d. %s" % (d["numero"], d["testo"])).font = \
            Font(name=FONT, bold=True, size=11)
        r += 1
        for testo, col in (("Opzione", 1), ("Conteggio", 2), ("%", 3)):
            c = wr.cell(row=r, column=col, value=testo)
            c.font = Font(name=FONT, bold=True, color="FFFFFF", size=10)
            c.fill = INTEST
        r += 1
        prima = r
        for o in d["opzioni"]:
            lettera = col_di[(d["id"], o["indice"])]
            wr.cell(row=r, column=1, value=o["etichetta"]).font = Font(name=FONT, size=10)
            wr.cell(row=r, column=2,
                    value="=SUM(Dati!%s2:%s%d)" % (lettera, lettera, max(2, ultima)))
            # percentuale sul numero di schede che hanno risposto alla domanda
            wr.cell(row=r, column=3,
                    value="=IFERROR(B%d/$B$2,0)" % r)
            for col in (2, 3):
                wr.cell(row=r, column=col).font = Font(name=FONT, size=10)
                wr.cell(row=r, column=col).border = BORDO
            wr.cell(row=r, column=3).number_format = "0.0%"
            wr.cell(row=r, column=1).border = BORDO
            r += 1
        if d["scala"]:
            lp = col_punteggio[d["id"]]
            wr.cell(row=r, column=1, value="Indice sintetico (1-5)").font = \
                Font(name=FONT, bold=True, size=10)
            wr.cell(row=r, column=2,
                    value="=IFERROR(AVERAGE(Dati!%s2:%s%d),\"\")" % (lp, lp, max(2, ultima)))
            wr.cell(row=r, column=2).number_format = "0.00"
            wr.cell(row=r, column=2).font = Font(name=FONT, bold=True, size=10)
            wr.cell(row=r, column=2).fill = GIALLO
            r += 1
        r += 1

    # indice complessivo
    wr.cell(row=r, column=1, value="INDICE COMPLESSIVO DI GRADIMENTO (media D%s)"
            % ", D".join("%02d" % d["numero"] for d in scala_domande)).font = \
        Font(name=FONT, bold=True, size=11, color="1F6FB2")
    medie = ["AVERAGE(Dati!%s2:%s%d)" % (col_punteggio[d["id"]],
                                         col_punteggio[d["id"]], max(2, ultima))
             for d in scala_domande]
    wr.cell(row=r, column=2, value="=IFERROR(AVERAGE(%s),\"\")" % ",".join(medie))
    wr.cell(row=r, column=2).number_format = "0.00"
    wr.cell(row=r, column=2).font = Font(name=FONT, bold=True, size=12)
    wr.cell(row=r, column=2).fill = GIALLO
    _larghezze(wr, [58, 14, 12])

    # ---------------- foglio Revisione ----------------
    wv = wb.create_sheet("Revisione")
    _intesta(wv, ["N.", "Pagina", "Cosa controllare", "Riferimento letto",
                  "Correzione manuale (scrivere qui)"])
    riga = 2
    for i, e in enumerate(esiti, start=1):
        if e["stato"] == STATO_OK:
            continue
        vals = [i, e["pagina"], " | ".join(e["note"]),
                e["payload"] or "(non letto)", ""]
        for j, v in enumerate(vals, start=1):
            c = wv.cell(row=riga, column=j, value=v)
            c.font = Font(name=FONT, size=10)
            c.border = BORDO
            c.alignment = Alignment(wrap_text=True, vertical="top")
        wv.cell(row=riga, column=5).fill = GIALLO
        riga += 1
    if riga == 2:
        c = wv.cell(row=2, column=1, value="Nessuna scheda da rivedere.")
        c.font = Font(name=FONT, italic=True, size=10)
    _larghezze(wv, [6, 22, 55, 40, 40])

    # ---------------- foglio Istruzioni ----------------
    wi = wb.create_sheet("Istruzioni")
    testi = [
        ("Come si legge questo file", True),
        ("", False),
        ("Risposte  \u2013 una riga per scheda, con il riferimento ricavato dal QR code.", False),
        ("Riepilogo \u2013 conteggi, percentuali e indice sintetico. Tutte formule: si", False),
        ("            aggiorna da solo se correggi qualcosa nel foglio Dati.", False),
        ("Revisione \u2013 le schede su cui il riconoscimento non e' sicuro. Vanno", False),
        ("            controllate a mano sulla carta e corrette nel foglio Dati.", False),
        ("Dati      \u2013 matrice 0/1 usata dalle formule. E' il foglio da correggere:", False),
        ("            metti 1 nella colonna dell'opzione giusta e 0 nelle altre.", False),
        ("            E' nascosto: tasto destro sulle linguette > Scopri.", False),
        ("", False),
        ("Indice sintetico: media 1-5 sulle domande di gradimento (1 = per niente,", False),
        ("5 = moltissimo). Permette di confrontare incontri, classi e annate.", False),
        ("", False),
        ("Le celle su sfondo giallo sono quelle da compilare o da tenere d'occhio.", False),
    ]
    if ha_commenti:
        testi[10:10] = [
            ("I commenti scritti a mano non vengono trascritti: l'ultima colonna", False),
            ("di Risposte contiene il percorso dell'immagine ritagliata.", False),
        ]
    for i, (t, grassetto) in enumerate(testi, start=1):
        c = wi.cell(row=i, column=1, value=t)
        c.font = Font(name=FONT, size=11 if grassetto else 10, bold=grassetto)
    _larghezze(wi, [95])

    wb.save(percorso)


def _col_index(lettera):
    n = 0
    for ch in lettera:
        n = n * 26 + (ord(ch) - 64)
    return n


# ==========================================================================
# 7. MAIN
# ==========================================================================
def main():
    ap = argparse.ArgumentParser(
        description="Legge le schede OMNIA-TFS scansionate e produce l'Excel.")
    ap.add_argument("input", nargs="+", help="PDF scansionati o immagini")
    ap.add_argument("--template", default="template.json",
                    help="Geometria prodotta da genera_schede.py")
    ap.add_argument("--out", default="valutazioni.xlsx", help="File Excel di uscita")
    ap.add_argument("--commenti", default="commenti",
                    help="Cartella dove salvare i ritagli dei commenti")
    ap.add_argument("--diagnostica", action="store_true",
                    help="Stampa i valori di riempimento di ogni casella")
    args = ap.parse_args()

    if not os.path.exists(args.template):
        print("Template %s non trovato. Generalo con:\n"
              "  python genera_schede.py --solo-template" % args.template)
        return 1
    with open(args.template, encoding="utf-8") as f:
        template = json.load(f)

    percorsi = []
    for p in args.input:
        espansi = glob.glob(p)
        percorsi.extend(espansi if espansi else [p])
    if not percorsi:
        print("Nessun file di input trovato.")
        return 1

    esiti = []
    for percorso in percorsi:
        try:
            pagine = pagine_da_file(percorso)
        except Exception as exc:
            print("  ! %s: %s" % (percorso, exc))
            continue
        for etichetta, gray in pagine:
            e = elabora_pagina(etichetta, gray, template, args.commenti,
                               len(esiti) + 1)
            esiti.append(e)
            simbolo = "." if e["stato"] == STATO_OK else "!"
            print("  %s %-24s %s" % (simbolo, etichetta,
                                     "" if e["stato"] == STATO_OK
                                     else " | ".join(e["note"])))
            if args.diagnostica:
                for d in template["domande"]:
                    print("      %s %s" % (d["id"], e["riempimenti"].get(d["id"])))

    if not esiti:
        print("Nessuna pagina elaborata.")
        return 1

    scrivi_excel(esiti, template, args.out)
    da_rivedere = sum(1 for e in esiti if e["stato"] != STATO_OK)
    print("\nSchede elaborate : %d" % len(esiti))
    print("Da rivedere      : %d (%.1f%%)"
          % (da_rivedere, 100.0 * da_rivedere / len(esiti)))
    print("File prodotto    : %s" % args.out)
    return 0


if __name__ == "__main__":
    sys.exit(main())
